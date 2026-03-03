import sys
from types import ModuleType

# Correção técnica para compatibilidade
if 'imghdr' not in sys.modules:
    sys.modules['imghdr'] = ModuleType('imghdr')

import streamlit as st
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from io import BytesIO

# --- CONFIGURAÇÃO DA MARCA NETMANIA ---
LOGO_NETMANIA = "https://netmaniainternet.com.br/img/logoSemFundo.png"

st.set_page_config(
    page_title="Netmania Optimizer", 
    page_icon=LOGO_NETMANIA, 
    layout="wide"
)

# --- CABEÇALHO COM LOGOTIPO ---
col_logo, col_titulo = st.columns([1, 4])
with col_logo:
    st.image(LOGO_NETMANIA, width=180) 
with col_titulo:
    st.title("Estruturador de Planilhas Personalizado")
    st.markdown("*Sistema de Gestão de Ativações e Reativações*")

# --- FUNÇÃO DE CARREGAMENTO INTELIGENTE ---
def carregar_dados_flexivel(arq, sem_header=False):
    if arq is None: return None
    header_val = None if sem_header else 0
    try:
        if arq.name.lower().endswith('.csv'):
            return pd.read_csv(arq, sep=None, engine='python', encoding='latin-1', header=header_val)
        return pd.read_excel(arq, header=header_val)
    except Exception as e:
        st.error(f"Erro ao ler arquivo {arq.name}: {e}")
        return None

# --- FUNÇÃO PARA LIMPAR DATAS ---
def formatar_apenas_data(valor):
    try:
        if pd.isna(valor) or valor == "" or valor == "REATIVAÇÃO": return valor
        dt = pd.to_datetime(valor)
        return dt.strftime('%d/%m/%Y')
    except:
        return str(valor)

# --- SEÇÃO DE UPLOAD ---
col1, col2, col3 = st.columns(3)
with col1: 
    arquivo_ativacao = st.file_uploader("1. Planilha de Ativação", type=['xlsx', 'csv', 'xlsm'])
    st.caption("⚠️ O sistema remove cancelados automaticamente.")

with col2: 
    arquivo_protocolos = st.file_uploader("2. Planilha de Protocolos", type=['xlsx', 'csv', 'xlsm'])
    st.caption("🔍 Filtros: Comercial Interno/Externo | Status: Abertura.")

with col3: 
    arquivo_reativacao = st.file_uploader("3. Relatório de Reativações", type=['xlsx', 'csv', 'xlsm'])
    st.caption("🔄 Filtros: Categoria 2 | Remover duplicados.")

# --- PROCESSAMENTO ---
if arquivo_ativacao:
    try:
        df_ativ = carregar_dados_flexivel(arquivo_ativacao)
        
        if df_ativ is not None:
            df_ativ.columns = [str(c).strip() for c in df_ativ.columns]

            # Filtro automático de contratos cancelados
            if 'Status Contrato' in df_ativ.columns:
                df_ativ = df_ativ[df_ativ['Status Contrato'].astype(str).str.lower() != 'cancelado']

            # --- INTEGRAÇÃO DE PROTOCOLOS (OPCIONAL) ---
            if arquivo_protocolos:
                df_prot = carregar_dados_flexivel(arquivo_protocolos)
                if df_prot is not None:
                    df_prot.columns = [str(c).strip() for c in df_prot.columns]
                    
                    col_vinc_ativ = 'Nome Cliente' if 'Nome Cliente' in df_ativ.columns else df_ativ.columns[6]
                    col_vinc_prot = 'Cliente' if 'Cliente' in df_prot.columns else df_prot.columns[15]

                    df_ativ['_JOIN'] = df_ativ[col_vinc_ativ].astype(str).str.strip().str.upper()
                    df_prot['_JOIN'] = df_prot[col_vinc_prot].astype(str).str.strip().str.upper()

                    col_resp = 'Responsavel' if 'Responsavel' in df_prot.columns else df_prot.columns[4]
                    df_prot_min = df_prot.drop_duplicates(subset=['_JOIN'])[['_JOIN', col_resp]]
                    
                    df_base = pd.merge(df_ativ, df_prot_min, on='_JOIN', how='left')
                    
                    if 'Responsavel' in df_base.columns and 'Vendedor 1' in df_base.columns:
                        df_base['Responsavel'] = df_base['Responsavel'].fillna(df_base['Vendedor 1'])
                    st.toast("Protocolos integrados!", icon="✅")
                else:
                    df_base = df_ativ.copy()
            else:
                df_base = df_ativ.copy()
                if 'Responsavel' not in df_base.columns and 'Vendedor 1' in df_base.columns:
                    df_base['Responsavel'] = df_base['Vendedor 1']
                st.warning("⚠️ **Aviso:** Sem planilha de Protocolos. O campo 'Responsável' foi preenchido com 'Vendedor 1'. Certifique-se de que os protocolos foram filtrados por *Comercial Interno/Externo* e *Status Abertura* caso decida subir o arquivo.")

            # --- INTEGRAÇÃO DE REATIVAÇÕES (OPCIONAL) ---
            if arquivo_reativacao:
                df_reat_raw = carregar_dados_flexivel(arquivo_reativacao, sem_header=True)
                if df_reat_raw is not None:
                    reat_rows = []
                    for _, row in df_reat_raw.iloc[1:].iterrows():
                        try:
                            reat_rows.append({
                                'Codigo Cliente': "REATIVAÇÃO",
                                'Contrato': row[35],
                                'Data Contrato': formatar_apenas_data(row[38]),
                                'Prazo Ativacao Contrato': formatar_apenas_data(row[8]),
                                'Ativacao Contrato': formatar_apenas_data(row[6]),
                                'Ativacao Conexao': formatar_apenas_data(row[10]),
                                'Nome Cliente': row[15],
                                'Responsavel': row[3],
                                'Vendedor 1': row[40],
                                'Endereco Ativacao': row[46],
                                'CEP': row[47],
                                'Cidade': row[44],
                                'Servico Ativado': row[42],
                                'Val Serv Ativado': "REATIVAÇÃO",
                                'Status Contrato': row[37],
                                'Assinatura Contrato': "",
                                'Vendedor 2': "",
                                'Origem': "REATIVAÇÃO",
                                'Valor Primeira Mensalidade': "REATIVAÇÃO"
                            })
                        except: continue
                    df_final_consolidado = pd.concat([df_base, pd.DataFrame(reat_rows)], ignore_index=True)
                    st.toast("Reativações incluídas!", icon="🔄")
                else:
                    df_final_consolidado = df_base
            else:
                df_final_consolidado = df_base
                st.info("ℹ️ **Nota:** Relatório de Reativações ausente. Apenas novas ativações serão exibidas. Lembre-se: para reativações, filtre por *Categoria 2* e remova duplicados.")

            # --- COLUNAS E FORMATAÇÃO ---
            colunas_finais = [
                'Codigo Cliente', 'Contrato', 'Data Contrato', 'Prazo Ativacao Contrato', 
                'Ativacao Contrato', 'Ativacao Conexao', 'Nome Cliente', 'Responsavel', 
                'Vendedor 1', 'Endereco Ativacao', 'CEP', 'Cidade', 'Servico Ativado', 
                'Val Serv Ativado', 'Status Contrato', 'Assinatura Contrato', 'Vendedor 2', 
                'Origem', 'Valor Primeira Mensalidade'
            ]
            
            df_export = df_final_consolidado[[c for c in colunas_finais if c in df_final_consolidado.columns]]
            
            for col in ['Data Contrato', 'Prazo Ativacao Contrato', 'Ativacao Contrato', 'Ativacao Conexao']:
                if col in df_export.columns:
                    df_export[col] = df_export[col].apply(formatar_apenas_data)

            st.dataframe(df_export, use_container_width=True)

            # --- EXCEL ESTILIZADO ---
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_export.to_excel(writer, index=False, sheet_name='Netmania')
                ws = writer.sheets['Netmania']
                
                # Estilos
                amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                verde = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
                vermelho_duplicado = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                fonte_corpo = Font(name='Calibri', size=11)
                centro = Alignment(horizontal='center', vertical='center')
                sem_grade = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

                for col_idx, col_cells in enumerate(ws.columns, 1):
                    header = ws.cell(row=1, column=col_idx)
                    if header.value == "Status Contrato": header.fill = amarelo
                    elif col_idx in [5, 15] or col_idx > 15: header.fill = verde
                    elif col_idx <= 9: header.fill = amarelo
                    
                    for cell in col_cells:
                        cell.font = fonte_corpo
                        cell.alignment = centro
                        cell.border = sem_grade
                    ws.column_dimensions[header.column_letter].width = 25

                # Duplicados
                max_row = ws.max_row
                for col_let in ['B', 'G']:
                    ws.conditional_formatting.add(
                        f'{col_let}2:{col_let}{max_row}',
                        FormulaRule(formula=[f'COUNTIF(${col_let}$2:${col_let}${max_row},{col_let}2)>1'], fill=vermelho_duplicado)
                    )

            st.success("✅ **Sucesso!** O arquivo está pronto para download.")
            st.download_button("📥 Baixar Planilha Final", output.getvalue(), "NETMANIA_CONSOLIDADO.xlsx")

    except Exception as e:
        st.error(f"Erro geral: {e}")

# --- RODAPÉ (TUTORIAL MANTIDO) ---
st.divider()
st.subheader("📖 Orientações Netmania")

t1, t2, t3 = st.columns(3)
with t1:
    st.markdown("### 📄 Ativação")
    st.info("Tratamento automático (o sistema já exclui os contratos com status 'Cancelado').")
with t2:
    st.markdown("### 📋 Protocolos")
    st.warning("Filtros necessários: **Comercial Interno/Externo** e Status: **Abertura**.")
with t3:
    st.markdown("### 🔄 Reativações")
    st.success("Filtros necessários: Remover duplicados e filtrar por **Categoria 2**.")
