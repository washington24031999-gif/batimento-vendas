import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO

# Configurações da página
st.set_page_config(page_title="Netmania Optimizer", layout="wide")
st.title("📊 Estruturador de Planilhas")

arquivo = st.file_uploader("Arraste sua planilha aqui", type=['xlsx', 'csv'])

if arquivo:
    try:
        # Leitura inteligente
        if arquivo.name.endswith('.csv'):
            df = pd.read_csv(arquivo, sep=None, engine='python', encoding='latin-1')
        else:
            df = pd.read_excel(arquivo)
            
        df.columns = [str(c).strip() for c in df.columns]

        # 1. Filtro de Cancelados
        if 'Status Contrato' in df.columns:
            df = df[df['Status Contrato'].str.lower() != 'cancelado']

        # 2. Ordem das colunas conforme suas imagens
        ordem = [
            'Codigo Cliente', 'Contrato', 'Data Contrato', 'Prazo Ativacao Contrato', 
            'Ativacao Contrato', 'Ativacao Conexao', 'Nome Cliente', 'Responsavel', 
            'Vendedor 1', 'Endereco Ativacao', 'CEP', 'Cidade', 'Servico Ativado', 
            'Val Serv Ativado', 'Status Contrato', 'Assinatura Contrato', 'Vendedor 2', 
            'Origem', 'Valor Primeira Mensalidade'
        ]
        
        df = df[[c for c in ordem if c in df.columns]]

        # 3. Formatação Visual em Memória
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Planilha')
            ws = writer.sheets['Planilha']
            
            # Estilos baseados nas fotos
            cor_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cor_verde = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
            fonte_calibri = Font(name='Calibri', size=11, bold=False)

            for col_idx, col_cells in enumerate(ws.columns, 1):
                header = ws.cell(row=1, column=col_idx)
                nome = str(header.value).strip()
                
                # Regras de Cores (Amarelo nas primeiras e no Status, Verde no final)
                if col_idx <= 9 or nome == "Status Contrato":
                    header.fill = cor_amarelo
                elif col_idx >= 16:
                    header.fill = cor_verde
                
                # Aplica Calibri 11 em tudo e ajusta largura das colunas
                for cell in col_cells:
                    cell.font = fonte_calibri
                ws.column_dimensions[header.column_letter].width = 22

        st.success("✅ Planilha processada com sucesso!")
        st.download_button(
            label="📥 Baixar Planilha Formatada",
            data=output.getvalue(),
            file_name="PLANILHA_NETMANIA_ORGANIZADA.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Erro ao processar o arquivo: {e}")